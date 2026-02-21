#!/usr/bin/env python3
"""
Build data/oews_metro_wage_distribution.json for the economic pillar wage-distribution signal.

BLS OEWS publishes metropolitan (and nonmetropolitan) area data with 10th, 25th, 75th, 90th
percentile wages. Download the annual metro area file from:
  https://www.bls.gov/oes/special-requests/oesm24ma.zip  (May 2024)
  https://www.bls.gov/oes/special-requests/oesm23ma.zip  (May 2023)

The zip contains XLSX files. This script expects either:
  - A path to the extracted directory or a single XLSX (--input),
  - Or it writes an empty JSON so the pillar still runs (wage_distribution will be null until populated).

To parse XLSX you need: pip install openpyxl

Run from project root:
  PYTHONPATH=. python3 scripts/build_oews_metro_wages.py
  PYTHONPATH=. python3 scripts/build_oews_metro_wages.py --input /path/to/oesm24ma
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(SCRIPT_DIR)
OUT_PATH = Path(ROOT) / "data" / "oews_metro_wage_distribution.json"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="", help="Path to OEWS metro zip extract dir or single XLSX.")
    ap.add_argument("--output", default=str(OUT_PATH), help="Output JSON path.")
    args = ap.parse_args()

    out: dict = {}

    if args.input and os.path.exists(args.input):
        path = Path(args.input)
        try:
            import openpyxl
        except ImportError:
            print("Install openpyxl to parse OEWS XLSX: pip install openpyxl", file=sys.stderr)
            _write_output(out, args.output)
            return

        # Find XLSX files (area codes often in filename or inside)
        if path.is_file() and path.suffix.lower() in (".xlsx", ".xls"):
            files = [path]
        else:
            files = list(path.rglob("*.xlsx")) or list(path.rglob("*.xls"))
        for f in files:
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                for sheet in wb.worksheets:
                    # BLS metro files: typically area code in col A or in a header row; 25th/75th columns
                    # Structure varies by release; adapt row/col if needed
                    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5000), values_only=True):
                        if not row or len(row) < 5:
                            continue
                        # Heuristic: first cell might be area code (e.g. 35620), look for numeric wage columns
                        area_code = None
                        p25 = p75 = None
                        for i, cell in enumerate(row):
                            if cell is None:
                                continue
                            if isinstance(cell, (int, float)):
                                if area_code is None and 10000 <= cell <= 99999 and i == 0:
                                    area_code = str(int(cell))
                                elif i > 5 and 10000 < cell < 500000:
                                    if p25 is None:
                                        p25 = cell
                                    elif p75 is None:
                                        p75 = cell
                                        break
                            elif isinstance(cell, str) and cell.strip().isdigit() and len(cell.strip()) == 5:
                                area_code = cell.strip()
                        if area_code and (p25 is not None or p75 is not None):
                            out[area_code] = {"wage_p25_annual": p25, "wage_p75_annual": p75}
                wb.close()
            except Exception as e:
                print(f"Skip {f}: {e}", file=sys.stderr)
    else:
        print("No --input path or file not found. Writing empty JSON.", file=sys.stderr)
        print("Download OEWS metro data from https://www.bls.gov/oes/special-requests/oesm24ma.zip", file=sys.stderr)

    _write_output(out, args.output)
    print(f"Wrote {len(out)} area(s) to {args.output}")


def _write_output(data: dict, path: str) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, sort_keys=True)


if __name__ == "__main__":
    main()
